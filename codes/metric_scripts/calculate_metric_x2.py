import os
import math
import numpy as np
import cv2
import glob
from sewar.full_ref import sam, uqi, scc
from openpyxl import Workbook, load_workbook

def main():
    # Configurations
    folder_GT = r''
    folder_Gen = r''

    crop_border = 4
    suffix = ''
    test_Y = False

    PSNR_all = []
    SSIM_all = []
    SAM_all = []
    QI_all = []
    SCC_all = []

    img_list = sorted(glob.glob(folder_GT + '\*'))

    if test_Y:
        print('Testing Y channel.')
    else:
        print('Testing RGB channels.')

    results = []

    for i, img_path in enumerate(img_list):
        base_name = os.path.splitext(os.path.basename(img_path))[0]
        im_GT = cv2.imread(img_path) / 255.
        im_Gen = cv2.imread(os.path.join(folder_Gen, base_name + suffix + '.tif')) / 255.

        if test_Y and im_GT.shape[2] == 3:
            im_GT_in = bgr2ycbcr(im_GT)
            im_Gen_in = bgr2ycbcr(im_Gen)
        else:
            im_GT_in = im_GT
            im_Gen_in = im_Gen

        if crop_border == 0:
            cropped_GT = im_GT_in
            cropped_Gen = im_Gen_in
        else:
            if im_GT_in.ndim == 3:
                cropped_GT = im_GT_in[crop_border:-crop_border, crop_border:-crop_border, :]
                cropped_Gen = im_Gen_in[crop_border:-crop_border, crop_border:-crop_border, :]
            elif im_GT_in.ndim == 2:
                cropped_GT = im_GT_in[crop_border:-crop_border, crop_border:-crop_border]
                cropped_Gen = im_Gen_in[crop_border:-crop_border, crop_border:-crop_border]
            else:
                raise ValueError('Wrong image dimension: {}. Should be 2 or 3.'.format(im_GT_in.ndim))

        PSNR = calculate_rgb_psnr(cropped_GT * 255, cropped_Gen * 255)
        SSIM = calculate_ssim(cropped_GT * 255, cropped_Gen * 255)
        SAM = sam(cropped_GT * 255, cropped_Gen * 255)
        QI = uqi(cropped_GT * 255, cropped_Gen * 255)
        SCC = scc(cropped_GT * 255, cropped_Gen * 255)

        print('{:3d} - {:25}. \tPSNR: {:.6f} dB, \tSSIM: {:.6f}, \tSAM: {:.6f}, \tQI: {:.6f}, \tSCC: {:.6f}'.format(
            i + 1, base_name, PSNR, SSIM, SAM, QI, SCC))
        
        PSNR_all.append(PSNR)
        SSIM_all.append(SSIM)
        SAM_all.append(SAM)
        QI_all.append(QI)
        SCC_all.append(SCC)
        
        results.append([base_name, PSNR, SSIM, SAM, QI, SCC])

    print('Average: PSNR: {:.6f} dB, SSIM: {:.6f}, SCC: {:.6f}, SAM: {:.6f}, QI: {:.6f}, '.format(
        sum(PSNR_all) / len(PSNR_all),
        sum(SSIM_all) / len(SSIM_all),
        sum(SCC_all) / len(SCC_all),
        sum(SAM_all) / len(SAM_all),
        sum(QI_all) / len(QI_all),  
    ))

    file_path = r"D:\Wz_Project_Learning\Super_Resolution_Reconstruction\HAUNet_RSISR\experiment\FSMambax2_UCMerced\x2_UCMerced.xlsx"
    save_data_to_excel(results, file_path)

def save_data_to_excel(results, file_name):
    try:
        try:
            wb = load_workbook(filename=file_name)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["File Name", "PSNR", "SSIM", "SAM", "QI", "SCC"])

        for result in results:
            ws.append(result)

        wb.save(file_name)
        print("数据已保存到Excel表格:", file_name)
    except Exception as e:
        print("保存数据到Excel表格时出错:", str(e))

def calculate_psnr(img1, img2):
    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    mse = np.mean((img1 - img2) ** 2)
    if mse == 0:
        return float('inf')
    return 20 * math.log10(255.0 / math.sqrt(mse))

def calculate_rgb_psnr(img1, img2):
    n_channels = np.ndim(img1)
    sum_psnr = 0
    for i in range(n_channels):
        this_psnr = calculate_psnr(img1[:, :, i], img2[:, :, i])
        sum_psnr += this_psnr
    return sum_psnr / n_channels

def ssim(img1, img2):
    C1 = (0.01 * 255) ** 2
    C2 = (0.03 * 255) ** 2

    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    kernel = cv2.getGaussianKernel(11, 1.5)
    window = np.outer(kernel, kernel.transpose())

    mu1 = cv2.filter2D(img1, -1, window)[5:-5, 5:-5]
    mu2 = cv2.filter2D(img2, -1, window)[5:-5, 5:-5]
    mu1_sq = mu1 ** 2
    mu2_sq = mu2 ** 2
    mu1_mu2 = mu1 * mu2
    sigma1_sq = cv2.filter2D(img1 ** 2, -1, window)[5:-5, 5:-5] - mu1_sq
    sigma2_sq = cv2.filter2D(img2 ** 2, -1, window)[5:-5, 5:-5] - mu2_sq
    sigma12 = cv2.filter2D(img1 * img2, -1, window)[5:-5, 5:-5] - mu1_mu2

    ssim_map = ((2 * mu1_mu2 + C1) * (2 * sigma12 + C2)) / ((mu1_sq + mu2_sq + C1) *
                                                            (sigma1_sq + sigma2_sq + C2))
    return ssim_map.mean()

def calculate_ssim(img1, img2):
    if not img1.shape == img2.shape:
        raise ValueError('Input images must have the same dimensions.')
    if img1.ndim == 2:
        return ssim(img1, img2)
    elif img1.ndim == 3:
        if img1.shape[2] == 3:
            ssims = []
            for i in range(img1.shape[2]):
                ssims.append(ssim(img1[..., i], img2[..., i]))
            return np.array(ssims).mean()
        elif img1.shape[2] == 1:
            return ssim(np.squeeze(img1), np.squeeze(img2))
    else:
        raise ValueError('Wrong input image dimensions.')

def bgr2ycbcr(img, only_y=True):
    in_img_type = img.dtype
    img.astype(np.float32)
    if in_img_type != np.uint8:
        img *= 255.
    if only_y:
        rlt = np.dot(img, [24.966, 128.553, 65.481]) / 255.0 + 16.0
    else:
        rlt = np.matmul(img, [[24.966, 112.0, -18.214], [128.553, -74.203, -93.786],
                              [65.481, -37.797, 112.0]]) / 255.0 + [16, 128, 128]
    if in_img_type == np.uint8:
        rlt = rlt.round()
    else:
        rlt /= 255.
    return rlt.astype(in_img_type)

if __name__ == '__main__':
    main()
